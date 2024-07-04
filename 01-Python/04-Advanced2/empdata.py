import openpyxl

wb = openpyxl.load_workbook("empData.xlsx")
emp_data = wb.active

def check_id(id = 0):
    if id == 0:
        id = input("Enter ID: ")
    max_rows = emp_data.max_row
    for i in range(2, max_rows + 1):
        if emp_data.cell(row=i,column=1).value == int(id):
            return i 
    return 0 

def add_data():
    # close excel if you gonna save
    id = input("Enter ID: ")
    while check_id(id):
        print("ID is taken")
        id = input("Enter ID: ")
    name = input("Enter Name: ")
    job = input("Enter Job: ")
    salary = input("Enter Salary: ")
    emp_data.append([int(id),name,job,salary])
    wb.save(filename="empData.xlsx")

    
def print_data(id):
    id_valid = check_id()
    if(not id_valid):
       print("This ID does not exist")
       return
    for i in range (1, emp_data.max_column + 1):
        print(emp_data.cell(row=id_valid,column= i).value,end = ",")

def remove_data(id):
    id_valid = check_id()
    if id_valid == 0:
        print("This ID does not exist") 
        return
    emp_data.delete_rows(id_valid)
    wb.save(filename="empData.xlsx")
    print("Empolyee data removed")


def update_data(id):
    id_valid = check_id()
    if id_valid == 0:
        print("This ID does not exist") 
        return
    ans = input("\nChoose the Value to modify (id/name/job/salary/exit): ")
    if ans.lower() == "id":
        emp_data.cell(row=id_valid,column=1).value = int(input("Enter New ID: "))
    elif ans.lower() == "name":
        emp_data.cell(row=id_valid,column=2).value = input("Enter New Name: ")
    elif ans.lower() == "job":
        emp_data.cell(row=id_valid,column=3).value = input("Enter New Job: ")
    elif ans.lower() == "salary":
        emp_data.cell(row=id_valid,column=4).value = input("Enter New Salary: ")
    else:
        return



while True:
    ans = input("\nChoose the Next action (add/print/remove/update/exit): ")
    
    if ans.lower() == "add":
        add_data()
    elif ans.lower() == "print":
        print_data(id)
    elif ans.lower() == "remove":
        remove_data(id)
    elif ans.lower() == "update":
        update_data(id)
    else:
        break

    
